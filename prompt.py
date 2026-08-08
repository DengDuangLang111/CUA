"""Prompt and schema for the spec generator. ROLE -> TASK -> RULES, then the
schema and one worked example carry the format.

Nothing here validates anything. The generator emits freely; quality is decided
downstream by the two host-side controls in emit.py and by the score==1.0
filter after the rollouts.
"""
import json

APPS = {
    "libreoffice_calc": ("libreoffice_calc", "LibreOffice Calc"),
    "libreoffice_writer": ("libreoffice_writer", "LibreOffice Writer"),
    "libreoffice_impress": ("libreoffice_impress", "LibreOffice Impress"),
    "chrome": ("chrome", "Google Chrome"),
    "gimp": ("gimp", None),
    "vlc": ("vlc", None),
    "thunderbird": ("thunderbird", None),
    "vscode": ("vscode", "Visual Studio Code"),
    "files": ("os", None),
    "terminal": ("os", None),
}

ARTIFACTS = [
    "spreadsheet", "text_document", "slide_deck", "raster_image", "media_file",
    "pdf_or_archive", "source_code", "terminal_output", "filesystem",
    "preference_store", "app_data_store", "email_store", "desktop_session",
    # 31 official tasks carry this artifact and 29 of them are graded on where
    # the browser ended up. It was missing from this list, so drawing one handed
    # the model an artifact its own schema rejected.
    "browser_tab",
]

SOURCES = ["self", "prompt_literal", "second_local_artifact", "live_web"]

# How a task is graded. Only `file` is graded by a probe we write; the other two
# hand the job to machinery OSWorld already ships, so for them there is nothing
# to write and nothing the host-side controls could check.
#
# The distinction matters because both controls run on the HOST, where there is
# no VM: they can only reach state solve_py can produce by writing files. The
# first batch that hit this produced vlc-set-loop-mode, whose probe checks
# ~/.config/vlc/vlcrc (a file solve_py wrote correctly) AND `pgrep -x vlc`. On the
# host both seed and solved print FAIL, so `positive` read "wrong:FAIL" and
# `negative` read "ok" without having tested anything. Marking the kind lets the
# report say "not applicable" instead of a failure that is not one, or -- worse --
# a green tick that verified nothing.
GOLD_KINDS = {
    "file": "probe_py reads files the agent created or edited. solve_py can "
            "reproduce the finished state on the host, so both controls are real.",
    "browser_state": "the answer is WHERE CHROME ENDED UP, not a file. Graded by "
                     "OSWorld's own url matcher against a list of regexes you "
                     "supply in url_patterns, so there is no probe_py and no "
                     "solve_py -- emit \"\" for both. Give start_url too.",
    "infeasible": "the task cannot be done in this environment and the correct "
                  "behaviour is to refuse. There is no probe and no solve_py: "
                  "OSWorld grades it from the agent's own FAIL signal.",
}

# Every non-file kind maps onto an evaluator shape OSWorld already ships, copied
# from what the official tasks do rather than reimplemented:
#
#   browser_state  func   is_expected_url_pattern_match      (chrome.py:71)
#                  result active_url_from_accessTree
#                  expect rule {"expected": [regex, ...]}    ALL must match
#                  -- 3 official chrome/multi_apps tasks are written exactly so
#   infeasible     func   infeasible, and nothing else       (27 official tasks,
#                  every one of them an evaluator with the single key "func")
#
# Deliberately NOT covered: the 8 needs_gui_only_state tasks. Official grades
# those with check_accessibility_tree, whose rules are CSS selectors or xpath
# over the GNOME accessibility XML. Writing those blind, for a desktop the
# generator never sees, is not something to ask a model to guess at.
EVALUATOR_SHAPE = {
    "browser_state": ("is_expected_url_pattern_match", "active_url_from_accessTree"),
    "infeasible": ("infeasible", None),
}

# Injected ahead of the generated programs, so the model never has to get the
# helper path or the root indirection right.
#
# P() is the guest home directory in ALL THREE programs. On the host it points at
# the build tree, which is laid out exactly like /home/user; in the VM it is
# /home/user itself. That single shared path vocabulary is what lets the same
# probe run as a host-side control and as the real grader. An earlier version had
# setup_py write into a flat directory and remapped basenames on the way out --
# it broke on every directory asset and on every target outside the uploaded set.
PROBE_PREAMBLE = '''import os, sys, traceback
TG_ROOT = os.environ.get("TG_ROOT", "/home/user")
sys.path.insert(0, os.environ.get("TG_HELP", "/tmp"))
def P(*a):
    return os.path.join(TG_ROOT, *a)
from tghelp import read_xlsx, read_docx, read_pptx, norm, num
def _tg_main():
'''

# The probe body is wrapped so that ANY exception becomes a clean FAIL on stdout.
# Without this a probe that raises prints nothing, get_vm_command_line hands the
# metric an empty string, and the task scores 0 -- correct for an idle agent, but
# indistinguishable from a correct agent hitting a probe bug. A probe that raises
# on the SOLVED state now fails the positive control instead of hiding.
PROBE_FOOTER = '''
try:
    _tg_main()
except Exception:
    traceback.print_exc(file=sys.stderr)
    print("FAIL")
'''


def build_probe(spec):
    """The exact probe source used both by the host controls and in the VM."""
    body = spec["probe_py"].rstrip("\n").splitlines() or ["pass"]
    indented = "\n".join(("    " + ln) if ln.strip() else ln for ln in body)
    return PROBE_PREAMBLE + indented + "\n" + PROBE_FOOTER

HOST_PREAMBLE = '''import os, sys
TG_ROOT = os.environ.get("TG_ROOT", ".")
def P(*a):
    p = os.path.join(TG_ROOT, *a)
    d = os.path.dirname(p)
    if d:
        os.makedirs(d, exist_ok=True)
    return p
'''

ROLE = """\
ROLE
You design computer-use tasks for an Ubuntu 22.04 GNOME desktop with LibreOffice
7.x, Chrome, GIMP, VLC, Thunderbird, VS Code, Files and a terminal. Every task is
graded automatically with no human in the loop, so a task is only worth writing
if a program can decide whether it was done.
"""

TASK = """\
TASK
Emit {n} task specs. Each one is four things:

  instruction  what the user wants, in plain English
  setup_py     writes the starting files. Everything it creates is copied into the
               VM before the agent starts.
  probe_py     runs INSIDE THE VM after the agent stops. It inspects the machine
               and prints exactly PASS or exactly FAIL. This is the answer key.
  solve_py     turns the starting files into the finished state a perfect agent
               would leave behind. It is a second, independent expression of
               "done"; if it and probe_py disagree the task is ill-specified and
               gets flagged.

In all three, P(...) is the user's home directory, so P("Desktop/sales.xlsx") is
the file the agent sees at /home/user/Desktop/sales.xlsx. Use the same path in
setup_py, solve_py and probe_py. Parent directories are created for you.

There is no gold file and no file comparison. probe_py decides, alone.

gold_kind says which of those four applies:
  file (default)  probe_py reads files; solve_py reproduces them. All four fields.
  browser_state   the answer is which page Chrome ended on. OSWorld grades it
                  against url_patterns; emit "" for probe_py and "" for solve_py.
  infeasible      the task cannot be done here and refusing IS the correct answer.
                  Emit "" for probe_py and "" for solve_py.
"""

RULES = """\
RULES
1  probe_py MUST print FAIL on the untouched setup files. A probe that can only
   print PASS gives every idle agent a perfect score, and nothing downstream can
   detect it. Make the check depend on work the agent has to do.
2  probe_py must check the SUBSTANCE of the task, not a proxy for it. If the task
   is "sort by date descending", check the whole ordering, not just the first row.
3  probe_py runs as the desktop user with stdlib, PIL, lxml and requests, plus the
   readers already imported for you: read_xlsx(p) -> {sheet: rows},
   read_docx(p) -> [paragraph], read_pptx(p) -> [[shape_text]], norm(v), num(v).
   There is NO openpyxl and NO python-docx in the VM. It gets 120 seconds, and no
   network except 127.0.0.1.
4  Live GUI state is readable and fair game: the accessibility tree of the whole
   desktop is at http://127.0.0.1:5000/accessibility, Chrome's tabs at
   http://127.0.0.1:1337/json, windows via `wmctrl -lx`.
5  setup_py and solve_py have stdlib plus openpyxl, docx, pptx and PIL, and they
   run on the host, so they cannot use anything only the VM has. No randomness,
   no clocks, no network. solve_py must reach the finished state by writing files,
   not by launching an application.
6  gold_kind=file (the default) means solve_py must be able to produce exactly the
   state probe_py checks. If the target is a config file the agent would create,
   solve_py writes it at the same path.
6b gold_kind=browser_state is for a task whose answer is which page Chrome ended
   up on. Emit "" for probe_py and "" for solve_py; OSWorld grades it itself.
   Give start_url (the page the agent starts from) and url_patterns, a list of
   REGEXES that the final URL must ALL match, e.g.
     ["^https://(www\\\\.)?dmv\\\\.virginia\\\\.gov/licenses-ids/license/applying/eligibility"]
   A single loose fragment like ["civil/documents-and-forms"] is also fine. Write
   the regex against the URL only -- page text is not available.
   Never make the answer depend on content that changes: no prices, no rankings,
   no availability, no "today's" anything. A page that is reachable by navigating
   a stable site structure is the target; put the argument in url_stability.
6c gold_kind=infeasible is for a task that CANNOT be done on this desktop, where
   the correct behaviour is to refuse. Emit "" for solve_py and "" for probe_py.
   The task must be genuinely impossible with the installed applications -- not
   merely tedious, not "needs a plugin you could install". State it as a normal,
   plausible-sounding user request; do not hint that it is impossible.
6e setup_shell runs commands INSIDE THE VM before the agent starts, as a list of
   SHELL STRINGS. They run as the desktop user, so anything system-wide needs
   sudo, and sudo needs the password piped in. Copy this form exactly -- it is
   what all six official install tasks use:
     ["echo {CLIENT_PASSWORD} | sudo -S apt-get update -y && echo {CLIENT_PASSWORD} | sudo -S apt-get install -y jq"]
   {CLIENT_PASSWORD} is substituted for you. A command that cannot succeed is not
   retried and not reported -- it silently does nothing and the agent then meets a
   machine missing the thing the task is about. Use setup_shell only when the task
   genuinely needs something installed or started.
7  The instruction must be complete on its own. Never point at a file for the real
   requirements. Never state the answer or a count of the answer.
8  Say what the user wants, never which control to operate. No "right-click", no
   "ctrl+h", no "use the terminal".
9  One unambiguous end state. If two careful people would produce different files,
   rewrite it until they would not.
10 Data at a realistic size for the scenario, roughly 10 to 40 rows or under two
   pages. Real-sounding names and values, not Item 1 / Company X.
11 Guest paths are lowercase with no spaces, given to P() without a leading slash.
12 If the agent edits a document in place, set save_target to that guest path so
   the file is flushed to disk before the probe runs.
13 Slugs are unique, lowercase, under 40 characters.
"""


def system_prompt():
    return ROLE + "\n" + RULES + "\nWorked example of one spec:\n\n" + json.dumps(
        EXAMPLE, ensure_ascii=False, indent=1) + "\n\nEmit every spec through the tool.\n"


# What each `operation` token asks the agent to do to the artifact. The label
# file carries the tokens but no glosses, and a bare token steers weakly -- with
# these one-liners the same (artifact, source) cell yields visibly different
# tasks depending on the verb, which is the whole point of adding the axis.
OPERATIONS = {
    "acquire": "locate information that is not present yet and bring it in",
    "derive": "compute new values from data already present",
    "rewrite": "replace existing content with corrected or restructured content",
    "add_element": "add something that is not there yet",
    "remove_element": "take something out",
    "set_value": "set a specific field, setting or property to a stated value",
    "restyle": "change appearance or formatting without changing the content",
    "re_encode": "produce the same content in a different format",
    "organize": "rearrange, group, rename or move things",
    "reach_state": "drive the machine or an application into a particular state",
}


def user_prompt(n, cells, priors=None):
    priors = priors or {}
    lines = []
    for i, c in enumerate(cells):
        op = c.get("operation") or ""
        extra = ""
        if c.get("needs_setup_shell"):
            extra += ", and it needs setup_shell to install or start something"
        if c.get("intent"):
            # Taxonomy brief: intent/domain/constraints lead, because they are
            # what the model has to invent around. artifact/primary follow as
            # constraints on what it may end in.
            lines.append(
                "  spec %d: intent=%s, domain=%s, constraints=%d, artifact=%s, "
                "source=%s, apps=%d, primary=%s%s"
                % (i + 1, c["intent"], c["domain"], c["constraints"],
                   c["artifact"], c["source"], c["app_count"], c["primary"], extra)
            )
        else:
            lines.append(
                "  spec %d: artifact=%s, source=%s, operation=%s, apps=%d, primary=%s, "
                "gold_kind=%s%s"
                % (i + 1, c["artifact"], c["source"], op or "any", c["app_count"],
                   c["primary"], c.get("gold_kind", "file"), extra)
            )

    kinds = sorted({c.get("gold_kind", "file") for c in cells})
    kind_text = ""
    if set(kinds) - {"file"}:
        kind_text = ("\n\ngold_kind for each spec is given above and is NOT yours to "
                     "change:\n" + "\n".join("  %s = %s" % (k, GOLD_KINDS[k])
                                             for k in kinds if k in GOLD_KINDS))

    tax = ""
    if any(c.get("intent") for c in cells):
        from ostg import taxonomy as T
        ints = sorted({c["intent"] for c in cells if c.get("intent")})
        doms = sorted({c["domain"] for c in cells if c.get("domain")})
        cons = sorted({c["constraints"] for c in cells if c.get("constraints")})
        tax = ("\n\nintent is what the user is fundamentally trying to get done:\n"
               + "\n".join("  %s = %s" % (i, T.INTENTS[i]) for i in ints)
               + "\n\ndomain is the business setting the scenario is dressed in "
                 "(%s). Invent realistic names, values and rules from that world -- "
                 "it is the main thing keeping two specs apart.\n"
                 % ", ".join(doms)
               + "\nconstraints is how many explicit requirements the instruction "
                 "imposes:\n"
               + "\n".join("  %d = %s" % (c, T.CONSTRAINTS[c]) for c in cons))

    ops = sorted({c.get("operation") for c in cells if c.get("operation")})
    op_text = ""
    if ops:
        op_text = ("\n\noperation is the kind of change the agent must make:\n"
                   + "\n".join("  %s = %s" % (o, OPERATIONS[o])
                               for o in ops if o in OPERATIONS))

    # Same cell, same four coordinates, every draw -- so without an explicit list
    # of what has already been written there the batches converge. Only slugs are
    # sent: they are compact and they are ours, not the official suite's.
    avoid = []
    for c in cells:
        done = priors.get((c["artifact"], c["source"])) or []
        if done:
            avoid.append("  %s / %s: %s"
                         % (c["artifact"], c["source"], ", ".join(sorted(set(done)))))
    avoid_text = ""
    if avoid:
        avoid_text = ("\n\nAlready generated in these cells. Do NOT repeat their "
                      "business scenario or their rule; go somewhere clearly "
                      "different:\n" + "\n".join(sorted(set(avoid))))

    return (
        TASK.format(n=n)
        + "\nPer-spec targets, in order:\n"
        + "\n".join(lines)
        + kind_text
        + tax
        + op_text
        + "\n\nartifact is what the probe inspects. source is where the information "
        "needed to do the task comes from: self = inside the artifact itself, "
        "prompt_literal = stated in the instruction, second_local_artifact = a "
        "different local file the agent must also open. Put the primary application "
        "first in apps."
        + avoid_text
        + "\n\nMake the batch diverse: different business domains, "
        "different kinds of work. Two specs must not share a business rule.\n"
    )


EXAMPLE = {
    "slug": "overtime-pay-rollup",
    "instruction": (
        "The workbook hours.xlsx on the Desktop lists each employee's department, "
        "hourly rate and hours worked for the week. Anything above 40 hours is "
        "overtime and is paid at 1.5 times the base rate. Fill in the Total Pay "
        "column for every employee, rounded to two decimals, and save the file."
    ),
    "artifact": "spreadsheet",
    "source": "self",
    "apps": ["libreoffice_calc"],
    "app_count": 1,
    "open_paths": ["Desktop/hours.xlsx"],
    "save_target": "Desktop/hours.xlsx",
    "setup_py": (
        "import openpyxl\n"
        "wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Hours'\n"
        "ws.append(['Name', 'Dept', 'Rate', 'Hours', 'Total Pay'])\n"
        "for r in [('Ana Reyes','Ops',22.5,44),('Ben Halvorsen','Ops',19.0,38),\n"
        "          ('Cy Okonkwo','Lab',31.25,52),('Dee Marchetti','Lab',27.0,40)]:\n"
        "    ws.append(list(r) + [None])\n"
        "wb.save(P('Desktop/hours.xlsx'))\n"
    ),
    "solve_py": (
        "import openpyxl\n"
        "wb = openpyxl.load_workbook(P('Desktop/hours.xlsx')); ws = wb['Hours']\n"
        "for row in ws.iter_rows(min_row=2):\n"
        "    rate, hours = row[2].value, row[3].value\n"
        "    pay = rate * min(hours, 40) + rate * 1.5 * max(hours - 40, 0)\n"
        "    row[4].value = round(pay, 2)\n"
        "wb.save(P('Desktop/hours.xlsx'))\n"
    ),
    "probe_py": (
        "rows = read_xlsx(P('Desktop/hours.xlsx'))['Hours']\n"
        "ok = len(rows) >= 5\n"
        "for r in rows[1:]:\n"
        "    rate, hours, got = num(r[2]), num(r[3]), num(r[4])\n"
        "    if rate is None or hours is None or got is None:\n"
        "        ok = False; break\n"
        "    want = rate * min(hours, 40) + rate * 1.5 * max(hours - 40, 0)\n"
        "    if abs(got - round(want, 2)) > 0.005:\n"
        "        ok = False; break\n"
        "print('PASS' if ok else 'FAIL')\n"
    ),
}

SCHEMA = {
    "type": "object",
    "required": ["specs"],
    "properties": {
        "specs": {
            "type": "array",
            "items": {
                "type": "object",
                "required": ["slug", "instruction", "artifact", "source", "apps",
                             "app_count", "setup_py", "solve_py", "probe_py"],
                "properties": {
                    "slug": {"type": "string", "description": "unique, lowercase, <40 chars"},
                    "gold_kind": {
                        "type": "string", "enum": list(GOLD_KINDS),
                        "description": "how the task is graded, default file. "
                                       + "; ".join("%s = %s" % kv for kv in GOLD_KINDS.items()),
                    },
                    "setup_shell": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "shell commands run INSIDE THE VM before the agent starts, "
                                       "as the desktop user. System changes need sudo with the "
                                       "password piped in: \"echo {CLIENT_PASSWORD} | sudo -S "
                                       "apt-get install -y jq\". Omit unless the task genuinely "
                                       "needs something installed or started.",
                    },
                    "start_url": {
                        "type": "string",
                        "description": "gold_kind=browser_state only: the page Chrome is opened "
                                       "at before the agent starts.",
                    },
                    "url_patterns": {
                        "type": "array", "items": {"type": "string"},
                        "description": "gold_kind=browser_state only: regexes the FINAL url must "
                                       "all match. re.search, so an unanchored fragment matches "
                                       "anywhere; escape dots you mean literally.",
                    },
                    "url_stability": {
                        "type": "string",
                        "description": "required when source=live_web: argue why this answer will "
                                       "not change. Acceptable: stable site structure, historical "
                                       "facts, archived pages, a URL pinned to a version or a "
                                       "Wikipedia oldid. Not acceptable: prices, rankings, "
                                       "availability, anything dated today.",
                    },
                    "instruction": {"type": "string", "description": "what the user wants; complete on its own; never states the answer"},
                    "artifact": {"type": "string", "enum": ARTIFACTS, "description": "what probe_py inspects"},
                    "source": {"type": "string", "enum": SOURCES},
                    "apps": {"type": "array", "items": {"type": "string", "enum": list(APPS)},
                             "description": "applications the agent must open and interact with, PRIMARY FIRST"},
                    "app_count": {"type": "integer", "minimum": 1, "maximum": 3},
                    "open_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "home-relative paths to open in their application before the agent starts, e.g. Desktop/sales.xlsx. Everything setup_py writes is uploaded regardless.",
                    },
                    "save_target": {"type": "string", "description": "home-relative path the agent edits in place, if any; flushed to disk before probing"},
                    "setup_py": {"type": "string", "description": "writes the starting files with P()"},
                    "solve_py": {"type": "string", "description": "turns the starting files into the finished state, in place, with P()"},
                    "probe_py": {"type": "string", "description": "VM program; prints exactly PASS or FAIL"},
                },
            },
        }
    },
}


def tool_definition(name="emit_task_specs"):
    return {
        "name": name,
        "description": "Emit the batch of task specs.",
        "input_schema": SCHEMA,
    }
